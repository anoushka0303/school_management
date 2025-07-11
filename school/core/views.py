from django.shortcuts import render
from rest_framework import viewsets, permissions, status, parsers
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from rest_framework_simplejwt.tokens import AccessToken, RefreshToken
from rest_framework.views import APIView
from .models import User, Student, Teacher, Principal, Course, Enrollment
from .serializers import *
from .permissions import *
from django.utils import timezone
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import FileResponse, Http404, HttpResponse
import os
from django.conf import settings
from .utils import send_email
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from io import BytesIO
from .pagination import *
from core import core_pb2 as core__pb2
from core import core_pb2_grpc as core__pb2__grpc

import grpc


User = get_user_model()
token_generator = PasswordResetTokenGenerator()
        
class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    pagination_class = UserDataPagination

    def get_permissions(self):
        if self.action == 'list':
            return [permissions.IsAuthenticated(), IsAdmin()]
        elif self.action == 'retrieve':
            return [permissions.IsAuthenticated(), IsAdminOrSelfForStudent()]
        elif self.action == 'destroy':
            return [permissions.IsAuthenticated(), IsAdmin()]
        return [permissions.IsAuthenticated()]
    

    def get_stub(self):
        channel = grpc.insecure_channel('localhost:50051')
        return core__pb2__grpc.StudentServiceStub(channel)
    
    def list(self, request, *args, **kwargs):
        if request.user.role != "admin":
            return Response({"error" : "only admin can list all students"}, status= status.HTTP_401_UNAUTHORIZED)
        grpc_stub = self.get_stub()
        grpc_request = core__pb2.GetStudentsRequest()
        grpc_response = grpc_stub.ListStudents(grpc_request)
        students_data = []
        for student in grpc_response.students:
            students_data.append({
                "student_id": student.student_id,
                "user_id": student.user_id.id,
                "name": student.name,
                "guardian_name": student.guardian_name,
                "guardian_contact": student.guardian_contact,
                "created_date": student.created_date.ToDatetime().isoformat() if student.HasField("created_date") else None,
                "class_name": student.class_name,
                "semester": student.semester,
            })

        page = self.paginate_queryset(students_data)
        if page is not None:
            return self.get_paginated_response(page)
        else:
            return Response(students_data)

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({"only admin can delete users."}, status=status.HTTP_403_FORBIDDEN)

        instance = self.get_object()
        instance.user.is_active = False
        instance.deleted_by = request.user
        instance.deleted_date = timezone.now()
        instance.created_by = request.user
        instance.created_date = timezone.now()
        instance.user.save()
        instance.save()
        return Response({"user has been deleted."}, status=status.HTTP_204_NO_CONTENT)


class TeacherViewSet(viewsets.ModelViewSet):
    queryset = Teacher.objects.all()
    serializer_class = TeacherSerializer
    pagination_class = UserDataPagination

    def get_permissions(self):
        if self.action == 'list':
            return [permissions.IsAuthenticated(), IsAdmin()]
        elif self.action == 'retrieve':
            return [permissions.IsAuthenticated(), IsAdminOrSelfForTeacher()]
        elif self.action == 'destroy':
            return [permissions.IsAuthenticated(), IsAdmin()]
        return [permissions.IsAuthenticated()]

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({"only admin can delete users."}, status=status.HTTP_403_FORBIDDEN)

        instance = self.get_object()
        instance.user.is_active = False
        instance.deleted_by = request.user
        instance.deleted_date = timezone.now()
        instance.created_by = request.user
        instance.created_date = timezone.now()
        instance.user.save()
        instance.save()
        return Response({"user has been deleted."}, status=status.HTTP_204_NO_CONTENT)


class PrincipalViewSet(viewsets.ModelViewSet):
    queryset = Principal.objects.all()
    serializer_class = PrincipalSerializer
    pagination_class = UserDataPagination

    def get_permissions(self):
        if self.action == 'list':
            return [permissions.IsAuthenticated(), IsAdmin()]
        elif self.action == 'retrieve':
            return [permissions.IsAuthenticated(), IsAdminOrSelfForPrincipal()]  
        elif self.action == 'destroy':
            return [permissions.IsAuthenticated(), IsAdmin()]
        return [permissions.IsAuthenticated()]

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({"only admin can delete users."}, status=status.HTTP_403_FORBIDDEN)

        instance = self.get_object()
        instance.user.is_active = False
        instance.deleted_by = request.user
        instance.deleted_date = timezone.now()
        instance.user.created_by = request.user
        instance.user.created_date = timezone.now()

        instance.created_by = request.user
        instance.created_date = timezone.now()
        instance.user.save()
        instance.save()
        return Response({"user has been deleted."}, status=status.HTTP_204_NO_CONTENT)


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = UserDataPagination


class EnrollmentViewSet(viewsets.ModelViewSet):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdmin]
    pagination_class = UserDataPagination


class GradeUpdateView(viewsets.ModelViewSet):
    queryset = Enrollment.objects.all()
    serializer_class = GradeUpdateSerializer
    permission_classes = [permissions.IsAuthenticated, CanUpdateOwnStudentGradeOnly]


class LoginView(viewsets.ViewSet):
    permission_classes = [AllowAny]

    def get_stub(self):
        channel = grpc.insecure_channel('localhost:50051')
        stub = core__pb2__grpc.AuthServiceStub(channel) 
        return stub, channel

    def grpc_post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')
        role = request.data.get('role')

        if not email or not password or not role:
            return Response({
                "message": "Bad request. Email, password and role needed",
                "status": 400,
                "has_error": True
            }, status=status.HTTP_400_BAD_REQUEST)

        stub, channel = self.get_stub()

        try:
            grpc_request = core__pb2.LoginRequest(
                email=email,
                password=password,
                role=role
            )
            #print(grpc_request)
            grpc_response = stub.Login(grpc_request)
            #print(grpc_response)

            user_data = {
                "id": grpc_response.user.id,
                "email": grpc_response.user.email,
                "role": grpc_response.user.role
            }
            '''if not user_data:
                print(user_data)
            else:
                print("1")'''

            profile_data = {}

            if grpc_response.HasField("student_profile"):
                profile_data = {
                    "student_id": grpc_response.student_profile.student_id,
                    "class_name": grpc_response.student_profile.class_name
                }
            elif grpc_response.HasField("teacher_profile"):
                profile_data = {
                    "teacher_id": grpc_response.teacher_profile.teacher_id
                }
            elif grpc_response.HasField("principal_profile"):
                profile_data = {
                    "principal_id": grpc_response.principal_profile.principal_id
                }
            '''else:
                return Response({
                    "error" : "invalid role"
                })'''

            response_data = {
                "access_token": grpc_response.access_token,
                "refresh_token": grpc_response.refresh_token,
                "user": user_data,
                "profile": profile_data
            }

            return Response(response_data, status=status.HTTP_200_OK)

        except grpc.RpcError as e:
            return Response({
                "message": f"gRPC error: {e.details()}",
                "status": 500,
                "has_error": True
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        finally:
            channel.close()

    def http_post(self, request):
        email = request.data.get('email')
        password = request.data.get('password')
        role = request.data.get('role')

        if not email or not password or not role:
            return Response({
                "message": f"Bad request. Email, password and role needed",
                "status" : 400,
                "has_error" : True
                }, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.filter(email=email).first()

        if not user or not user.check_password(password):
            return Response({
                "message": f"Unauthorized access. Invalid credentials.",
                "status" : 401,
                "has_error" : True,
                }, status=401)

        if user.role != role:
            return Response({
                "message": f"Unauthorized access. User role mismatch. Expected role '{user.role}'.",
                "status" : 401,
                "has_error" : True,
                }, status=401)

        access_token = AccessToken.for_user(user)
        refresh_token = RefreshToken.for_user(user)
        return Response({
            "messgae" : "Login Successful",
            "has_error" : False,
            "status" : 200,
            "access": str(access_token),
            "refresh" : str(refresh_token),
            "user": {
                "email": user.email,
                "role": user.role,
            }
        }, status=200)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrSelf]
    pagination_class = UserDataPagination

    def get_queryset(self):
        user = self.request.user
        if user.role == 'admin':
            return User.objects.all()
        return User.objects.filter(id=user.id)


class Register(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def get_stub(self):
        channel = grpc.insecure_channel('localhost:50051')
        stub = core__pb2__grpc.AuthServiceStub(channel) 
        return stub, channel

    def http_post(self, request):
        role = request.data.get('role')
        email = request.data.get('email')
        password = request.data.get('password')
        profile_data = request.data.get('profile', {})

        if not all([email, password, role]):
            return Response({
                "has_error" : True,
                "status" : 400,
                "message": "Bad Request. Email, password, and role are required"
                }, status=400)

        if role not in ['student', 'teacher', 'principal']:
            return Response({
                "has_error" : True,
                "status" : 400,
                "message": "Bad Request. Only student, teacher or principal can be created by admin"
                }, status=400)

        user = User.objects.create_user(email=email, password=password, role=role, created_date = timezone.now())
        profile_data['user'] = user.id

        if role == 'student':
            serializer = StudentSerializer(data=profile_data)
        elif role == 'teacher':
            serializer = TeacherSerializer(data=profile_data)
        elif role == 'principal':
            serializer = PrincipalSerializer(data=profile_data)

        if serializer.is_valid():
            instance = serializer.save()
            instance.created_by = request.user
            instance.created_date = timezone.now()
            instance.user.created_by = request.user
            instance.user.created_date = timezone.now()
            instance.save()

            admin_email = request.user.email
            subject = "You have been registered on the platform"
            html = f"""
            <p>Hello,</p>
            <p>You have been registered as a <strong>{role}</strong> on the platform by admin <strong>{admin_email}</strong>.</p>
            <p>Login email: <strong>{email}</strong></p>
            <p>Password : {password}</p>
            <p>Please reset your password after logging in.</p>
            """

            send_email(to_email=email,subject= subject,html_content= html)

            return Response({
                "user": UserSerializer(user).data,
                "profile": serializer.data
            })
        else:
            user.delete()
            return Response(serializer.errors, status=400)
        

    def grpc_post(self, request):
        stub, channel = self.get_stub()

        access_token = str(request.auth)  

        metadata = [('authorization', f'Bearer {access_token}')]
        if not access_token:
            return Response({
                "message": "Missing Authorization header",
                "status": 401,
                "has_error": True
            }, status=401)


        email = request.data.get('email')
        password = request.data.get('password')
        role = request.data.get('role')

        if not all([email, password, role]):
            return Response({
                "message": "Email, password and role are required.",
                "status": 400,
                "has_error": True
            }, status=400)

        try:
            grpc_request = core__pb2.RegisterRequest(
                email=email,
                password=password,
                role=role
            )

            if role.lower() == "student":
                student_data = request.data.get('profile', {})
                student_profile = core__pb2.Student(
                    name=student_data.get('name'),
                    guardian_name=student_data.get('guardian_name'),
                    guardian_contact=student_data.get('guardian_contact'),
                    student_contact=student_data.get('student_contact'),
                    class_name=student_data.get('class_name'),
                    semester=student_data.get('semester', 1),
                    address=student_data.get('address'),
                    fee_status=student_data.get('fee_status')
                )
                course_ids = student_data.get('courses')
                if not course_ids:
                    print("none")
                    course_ids = []
                else:
                    print(f"course_ids: {course_ids} type: {type(course_ids)}")
                student_profile.course_ids.extend(course_ids)
                grpc_request.student_profile.CopyFrom(student_profile)

            elif role == "teacher":
                grpc_request.teacher_profile.CopyFrom(core__pb2.Teacher())
            elif role == "principal":
                grpc_request.principal_profile.CopyFrom(core__pb2.Principal())

            grpc_response = stub.Register(grpc_request, metadata=metadata)

            user_data = {
                "id": grpc_response.user.id,
                "email": grpc_response.user.email,
                "role": grpc_response.user.role
            }

            profile_data = {}

            if grpc_response.HasField("student_profile"):
                profile_data = {
                    "student_id": grpc_response.student_profile.student_id,
                    "name": grpc_response.student_profile.name,
                    "guardian_name": grpc_response.student_profile.guardian_name,
                    "guardian_contact": grpc_response.student_profile.guardian_contact,
                    "student_contact": grpc_response.student_profile.student_contact,
                    "course_ids": list(grpc_response.student_profile.course_ids),
                    "class_name": grpc_response.student_profile.class_name,
                    "semester": grpc_response.student_profile.semester,
                    "created_date": str(grpc_response.student_profile.created_date),
                    "created_by": int(grpc_response.student_profile.created_by),
                    "updated_date": str(grpc_response.student_profile.updated_date),
                    "updated_by": grpc_response.student_profile.updated_by,
                    "deleted_date": str(grpc_response.student_profile.deleted_date),
                    "deleted_by": grpc_response.student_profile.deleted_by,
                }
            elif grpc_response.HasField("teacher_profile"):
                profile_data = {
                    "teacher_id": grpc_response.teacher_profile.teacher_id
                }
            elif grpc_response.HasField("principal_profile"):
                profile_data = {
                    "principal_id": grpc_response.principal_profile.principal_id
                }

            response_data = {
                "message": "Registration successful",
                "has_error": False,
                "user": user_data,
                "profile": profile_data
            }
            return Response(response_data, status=201)

        except grpc.RpcError as e:
            return Response({
                "message": f"gRPC error: {e.details()}",
                "status": 500,
                "has_error": True
            }, status=500)

        finally:
            channel.close()


    
    
class RequestPasswordResetView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        email = request.data.get('email')
        old_password = request.data.get('old_password')

        if not email or not old_password:
            return Response({
                "has_error" : True,
                "status" : 400,
                "message": "Bad request. Email and previous password is required"
                }, status=400)

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return Response({
                "has_error" : True,
                "status" : 404,
                "message": "User not found"
                }, status=404)
        
        if not user.is_active:
            return Response({
                "has_error" : True,
                "status" : 404,
                "message": "User not found"
                }, status= 404)
        
        if not user.check_password(old_password):
            return Response({
                "has_error" : True,
                "status" : 403,
                "message" : "Forbidden error. Old password does not match."
                }, status= 403)

        uid = urlsafe_base64_encode(force_bytes(user.pk))  
        token = PasswordResetTokenGenerator().make_token(user)  

        return Response({
            "uid": uid,
            "token": token
        }, status=200)
    

class ConfirmPasswordResetView(APIView):
    permission_classes = [AllowAny]
    def post(self, request):
        uid = request.data.get('uid')
        token = request.data.get('token')
        new_password = request.data.get('new_password')

        if not all([uid, token, new_password]):
            return Response({"error": "uid, token and new_password are required"}, status=400)

        try:
            uid = force_str(urlsafe_base64_decode(uid))
            user = User.objects.get(pk=uid)
        except (User.DoesNotExist, ValueError, TypeError):
            return Response({"error": "Invalid UID"}, status=400)

        if not token_generator.check_token(user, token):
            return Response({"error": "Invalid or expired token"}, status=400)

        user.set_password(new_password)
        user.save()
        return Response({"message": "Password reset successful"}, status=200)

    
class StudentUpdateView(viewsets.ViewSet):
    serializer_class = StudentUpdateSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrSelfForStudent]

    def get_object(self):
        if self.request.user.is_staff:
            student_id = self.kwargs.get('pk')
            return Student.objects.get(pk=student_id)
        return Student.objects.get(user=self.request.user)

    def partial_update(self, request, pk=None):
        instance = self.get_object()
        self.check_object_permissions(request, instance)
        serializer = self.serializer_class(instance, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class TeacherUpdateView(viewsets.ModelViewSet):
    serializer_class = TeacherUpdateSerializer
    permission_classes = [permissions.IsAuthenticated, IsTeacher, IsAdminOrSelfForTeacher]

    def get_object(self):
        return Teacher.objects.get(user=self.request.user)


class PrincipalUpdateView(viewsets.ModelViewSet):
    serializer_class = PrincipalUpdateSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrSelfForPrincipal]
    http_method_names = ['patch']  

    def get_object(self):
        return Principal.objects.get(user=self.request.user)

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.check_object_permissions(request, instance)
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)
    

class UpdateFeeStatusView(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentFeeStatusSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdmin]

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        self.check_object_permissions(request, instance) 

        serializer = self.get_serializer(instance, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)


class BulkDownloadExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated, IsAdmin]

    def get(self, request, role):
        if role == 'student':
            ids = request.query_params.get('id')
            queryset = Student.objects.filter(student_id__in=[int(i) for i in ids.split(',') if i.isdigit()]) if ids else Student.objects.all()
            serializer = StudentSerializer(queryset, many=True)
            sheet_name = "Students"
            file_name = "students.xlsx"

        elif role == 'teacher':
            ids = request.query_params.get('id')
            queryset = Teacher.objects.filter(faculty_id__in=[int(i) for i in ids.split(',') if i.isdigit()]) if ids else Teacher.objects.all()
            serializer = TeacherSerializer(queryset, many=True)
            sheet_name = "Teachers"
            file_name = "teachers.xlsx"

        elif role == 'principal':
            ids = request.query_params.get('id')
            queryset = Principal.objects.filter(principal_id__in=[int(i) for i in ids.split(',') if i.isdigit()]) if ids else Principal.objects.all()
            serializer = PrincipalSerializer(queryset, many=True)
            sheet_name = "Principals"
            file_name = "principals.xlsx"

        else:
            return Response({"error": "Invalid role"}, status=400)


        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = list(serializer.data[0].keys()) if serializer.data else []
        ws.append(headers)

        for data in serializer.data:
            row = []
            for field in headers:
                value = data.get(field, "")
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif isinstance(value, dict):
                    value = str(value)
                row.append(value)
            ws.append(row)


        folder_path = os.path.join(settings.MEDIA_ROOT, 'exports')
        os.makedirs(folder_path, exist_ok=True)

        file_path = os.path.join(folder_path, file_name)
        wb.save(file_path)

        file_url = request.build_absolute_uri(settings.MEDIA_URL + 'exports/' + file_name)
        return Response({'file_url': file_url})
    

class BulkUploadViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated, IsAdmin]

    def create(self, request):
        users_data = request.data.get("users", [])
        results = []
        for user_data in users_data:
            email = user_data.get("email")
            password = user_data.get("password")
            role = user_data.get("role")
            profile_data = user_data.get("profile", {})

            if not all([email, password, role]):
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": "Missing email, password, or role"
                })
                continue

            if role not in ['student', 'teacher', 'principal']:
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": "Invalid role"
                })
                continue

            try:
                user = User.objects.create_user(
                    email=email,
                    password=password,
                    role=role,
                    created_date=timezone.now()
                )
                profile_data["user"] = user.id

                if role == 'student':
                    serializer = StudentSerializer(data=profile_data)
                elif role == 'teacher':
                    serializer = TeacherSerializer(data=profile_data)
                elif role == 'principal':
                    serializer = PrincipalSerializer(data=profile_data)

                if serializer.is_valid():
                    instance = serializer.save()
                    instance.created_by = request.user
                    instance.created_date = timezone.now()
                    instance.user.created_by = request.user
                    instance.user.created_date = timezone.now()
                    instance.save()

                    # Email
                    subject = "You have been registered on the platform"
                    html = f"""
                    <p>Hello,</p>
                    <p>You have been registered as a <strong>{role}</strong> on the platform by admin <strong>{request.user.email}</strong>.</p>
                    <p>Login email: <strong>{email}</strong></p>
                    <p>Password : {password}</p>
                    <p>Please reset your password after logging in.</p>
                    """
                    send_email(to_email=email, subject=subject, html_content=html)

                    results.append({
                        "email": email,
                        "status": "success",
                        "user": UserSerializer(user).data,
                        "profile": serializer.data
                    })
                else:
                    user.delete()
                    results.append({
                        "email": email,
                        "status": "failed",
                        "message": serializer.errors
                    })
            except Exception as e:
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": str(e)
                })

        return Response({"results": results})


class BulkEnrollViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated, IsAdmin]

    def create(self, request):
        users_data = request.data.get("enrollments", [])
        results = []

        for data in users_data:
            course = data.get("course")
            student = data.get("student")
            grade = data.get("grade", None)

            if not course or not student:
                results.append({
                    "status": "failed",
                    "student": student,
                    "course": course,
                    "message": "Student or course missing"
                })
                continue

            serializer = EnrollmentSerializer(data={
                "student": student,
                "course": course,
                "grade": grade
            })

            if serializer.is_valid():
                serializer.save()
                results.append({
                    "status": "success",
                    "data": serializer.data
                })
            else:
                results.append({
                    "status": "failed",
                    "errors": serializer.errors
                })

        return Response(results)
     

class UploadExcel(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated, IsAdmin]
    parser_classes = [parsers.MultiPartParser]
    

    def create(self, request):
        #print("Entering create method")
        excel = request.FILES.get("file")
        '''if not excel:
            print("none")
        else:
            print(excel)'''

        if not excel:
            return Response({
                "message": "Bad Request. Excel file expected.",
                "has_error": True,
                "status": 400
            }, status=400)

        try:
            file_name = excel.name.lower()
            relative_path = f"bulk_uploads/{excel.name}"
            file_instance = default_storage.save(relative_path, ContentFile(excel.read()))
            #print(excel.read())
            #print(file_instance)
            absolute_path = default_storage.path(file_instance)

            status_obj = BulkUploadStatus.objects.create(
                uploaded_by=request.user,
                file_name=file_name,
                status='uploading',
                total_record=0,     
                success_count=0,
                failure_count=0,
                file = absolute_path
            )
            if "student" in file_name:
                return self.UploadStudents(excel, request, status_obj)
            elif "teacher" in file_name:
                return self.UploadTeachers(excel, request, status_obj)
            elif "principal" in file_name:
                return self.UploadPrincipals(excel, request, status_obj)
            else:
                return Response({
                    "message" : "Invalid file type. Must contain 'teacher', 'student' or 'principal' in file name.",
                    "status" : 400
                }, status= 400)
        except Exception as e:
            return Response({
                "error": str(e),
                "status": 500
            }, status=500)
        
    def UploadStudents(self, excel_file, request, status_obj):
        success_count = 0
        fail_count = 0
        total = 0
        serializer_class = StudentSerializer
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]

        results = []

        '''required_fields = ['email', 'password', 'role']
        if not all(field in headers for field in required_fields):
            fail_count += 1
            results.append({
                "message": "Invalid data in file. Email, password and role columns are required.",
                "status": 400
            })'''

        for row in sheet.iter_rows(min_row=2, values_only=True):
            #print(row)
            total += 1
            row_data = dict(zip(headers, row))
            email = row_data.get('email')
            password = row_data.get('password')
            role = row_data.get('role')

            if not email or not password or not role:
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": "Missing email, password, or role."
                })
                continue

            if role != "student":
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": "This upload only supports role=student"
                })
                continue

            profile_data = {}
            course_ids = []

            for key, value in row_data.items():
                if key in ['email', 'password', 'role']:
                    continue
                elif key == "courses" and value:
                    course_ids = [int(cid.strip()) for cid in str(value).split(",") if cid.strip().isdigit()]
                    profile_data[key] = course_ids
                else:
                    profile_data[key] = value

            try:
                user = User.objects.create_user(
                    email=email,
                    password=password,
                    role=role,
                    created_date=timezone.now()
                )
            except Exception as e:
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": str(e)
                })
                continue

            profile_data['user'] = user.id
            serializer = serializer_class(data=profile_data)

            if serializer.is_valid():
                success_count += 1
                instance = serializer.save()  
                for cid in course_ids:
                    try:
                        course = Course.objects.get(course_id=cid)
                        Enrollment.objects.get_or_create(student=instance, course=course)
                    except Course.DoesNotExist:
                        results.append({
                            "email": email,
                            "status": "warning",
                            "message": f"Course ID {cid} not found"
                        })

                instance.created_by = request.user
                instance.created_date = timezone.now()
                instance.user.created_by = request.user
                instance.user.created_date = timezone.now()
                instance.save()

                subject = "You have been registered on the platform"
                html = f"""
                <p>Hello,</p>
                <p>You have been registered as a <strong>{role}</strong> on the platform by admin <strong>{request.user.email}</strong>.</p>
                <p>Login email: <strong>{email}</strong></p>
                <p>Password : {password}</p>
                <p>Please reset your password after logging in.</p>
                """
                send_email(to_email=email, subject=subject, html_content=html)

                results.append({
                    "email": email,
                    "status": "success",
                    "data": serializer.data
                })
            else:
                fail_count += 1
                user.delete()
                results.append({
                    "email": email,
                    "status": "failed",
                    "errors": serializer.errors
                })
        status_obj.total_record = total
        status_obj.success_count = success_count
        status_obj.failure_count = fail_count

        if fail_count == total:
            status_obj.status = 'fail'
        elif success_count == total:
            status_obj.status = 'success'
        else:
            status_obj.status = 'partial_success'


        wb = openpyxl.load_workbook(status_obj.file)
        sheet = wb.active
        existing_headers = [cell.value for cell in sheet[1]]


        if "Status" not in existing_headers:
            sheet.cell(row=1, column=len(existing_headers) + 1).value = "Status"


        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=0):
            result = results[i] if i < len(results) else {}
            msg = result.get("message") or str(result.get("errors") or result.get("status", "unknown"))
            sheet.cell(row=i + 2, column=len(existing_headers) + 1).value = msg

        wb.save(status_obj.file)
        status_obj.save()

        return Response({"results": results}, status=200)
    

    def UploadTeachers(self, excel_file, request, status_obj):
        total = 0
        fail_count = 0
        success_count = 0
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        serializer_class = TeacherSerializer
        headers = [cell.value for cell in sheet[1]]

        '''required_fields = ['email', 'password', 'role']
        if not all(field in headers for field in required_fields):
            return Response({
                "message": "Invalid data in file. Email, password and role columns are required.",
                "status": 400
            }, status=400)'''

        results = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            total += 1
            row_data = dict(zip(headers, row))
            email = row_data.get('email')
            password = row_data.get('password')
            role = row_data.get('role')

            if not email or not password or not role:
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": "Missing email, password, or role."
                })
                continue

            if role != "teacher":
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": "This upload only supports role=teacher"
                })
                continue

            profile_data = {}

            for key, value in row_data.items():
                if key in ['email', 'password', 'role']:
                    continue
                else:
                    profile_data[key] = value

            try:
                user = User.objects.create_user(
                    email=email,
                    password=password,
                    role=role,
                    created_date=timezone.now()
                )
            except Exception as e:
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": str(e)
                })
                continue

            profile_data['user'] = user.id
            serializer = serializer_class(data=profile_data)

            if serializer.is_valid():
                success_count += 1
                instance = serializer.save()  
                instance.created_by = request.user
                instance.created_date = timezone.now()
                instance.user.created_by = request.user
                instance.user.created_date = timezone.now()
                instance.save()

                subject = "You have been registered on the platform"
                html = f"""
                <p>Hello,</p>
                <p>You have been registered as a <strong>{role}</strong> on the platform by admin <strong>{request.user.email}</strong>.</p>
                <p>Login email: <strong>{email}</strong></p>
                <p>Password : {password}</p>
                <p>Please reset your password after logging in.</p>
                """
                send_email(to_email=email, subject=subject, html_content=html)

                results.append({
                    "email": email,
                    "status": "success",
                    "data": serializer.data
                })
            else:
                fail_count += 1
                user.delete()
                results.append({
                    "email": email,
                    "status": "failed",
                    "errors": serializer.errors
                })

        status_obj.total_record = total
        status_obj.success_count = success_count
        status_obj.failure_count = fail_count

        if fail_count == total:
            status_obj.status = 'fail'
        elif success_count == total:
            status_obj.status = 'success'
        else:
            status_obj.status = 'partial_success'

        wb = openpyxl.load_workbook(filename=default_storage.path(status_obj.file))
        sheet = wb.active
        existing_headers = [cell.value for cell in sheet[1]]


        if "Status" not in existing_headers:
            sheet.cell(row=1, column=len(existing_headers) + 1).value = "Status"

        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=0):
            result = results[i] if i < len(results) else {}
            msg = result.get("message") or str(result.get("errors") or result.get("status", "unknown"))
            sheet.cell(row=i + 2, column=len(existing_headers) + 1).value = msg

        wb.save(default_storage.path(status_obj.file))
        status_obj.save()

        return Response({"results": results}, status=200)
    
    def UploadPrincipals(self, excel_file, request, status_obj):
        total = 0
        success_count = 0
        fail_count = 0
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        serializer_class = PrincipalSerializer
        headers = [cell.value for cell in sheet[1]]

        '''required_fields = ['email', 'password', 'role']
        if not all(field in headers for field in required_fields):
            return Response({
                "message": "Invalid data in file. Email, password and role columns are required.",
                "status": 400
            }, status=400)'''

        results = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            total += 1
            row_data = dict(zip(headers, row))
            email = row_data.get('email')
            password = row_data.get('password')
            role = row_data.get('role')

            if not email or not password or not role:
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": "Missing email, password, or role."
                })
                continue

            if role != "principal":
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": "This upload only supports role=principal"
                })
                continue

            profile_data = {}

            for key, value in row_data.items():
                if key in ['email', 'password', 'role']:
                    continue
                else:
                    profile_data[key] = value

            try:
                user = User.objects.create_user(
                    email=email,
                    password=password,
                    role=role,
                    created_date=timezone.now()
                )
            except Exception as e:
                fail_count += 1
                results.append({
                    "email": email,
                    "status": "failed",
                    "message": str(e)
                })
                continue

            profile_data['user'] = user.id
            serializer = serializer_class(data=profile_data)

            if serializer.is_valid():
                success_count += 1
                instance = serializer.save()  
                instance.created_by = request.user
                instance.created_date = timezone.now()
                instance.user.created_by = request.user
                instance.user.created_date = timezone.now()
                instance.save()

                subject = "You have been registered on the platform"
                html = f"""
                <p>Hello,</p>
                <p>You have been registered as a <strong>{role}</strong> on the platform by admin <strong>{request.user.email}</strong>.</p>
                <p>Login email: <strong>{email}</strong></p>
                <p>Password : {password}</p>
                <p>Please reset your password after logging in.</p>
                """
                send_email(to_email=email, subject=subject, html_content=html)

                results.append({
                    "email": email,
                    "status": "success",
                    "data": serializer.data
                })
            else:
                fail_count += 1
                user.delete()
                results.append({
                    "email": email,
                    "status": "failed",
                    "errors": serializer.errors
                })

        status_obj.total_record = total
        status_obj.success_count = success_count
        status_obj.failure_count = fail_count

        if fail_count == total:
            status_obj.status = 'fail'
        elif success_count == total:
            status_obj.status = 'success'
        else:
            status_obj.status = 'partial_success'

        wb = openpyxl.load_workbook(filename=default_storage.path(status_obj.file))
        sheet = wb.active
        existing_headers = [cell.value for cell in sheet[1]]


        if "Status" not in existing_headers:
            sheet.cell(row=1, column=len(existing_headers) + 1).value = "Status"


        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=0):
            result = results[i] if i < len(results) else {}
            msg = result.get("message") or str(result.get("errors") or result.get("status", "unknown"))
            sheet.cell(row=i + 2, column=len(existing_headers) + 1).value = msg

        wb.save(default_storage.path(status_obj.file))
        status_obj.save()

        return Response({"results": results}, status=200)


class DownloadResultViewSet(APIView):
    permission_classes = [IsAdmin, permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            status_obj = BulkUploadStatus.objects.get(pk=pk)
            abs_path = status_obj.file  

            if not os.path.exists(abs_path):
                raise FileNotFoundError

            return FileResponse(
                open(abs_path, 'rb'),
                as_attachment=True,
                filename=os.path.basename(abs_path)
            )

        except BulkUploadStatus.DoesNotExist:
            raise Http404("Upload status not found")

        except FileNotFoundError:
            return Response({"error": "File not found on server"}, status=404)